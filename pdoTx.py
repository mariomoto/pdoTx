# -*- coding: utf-8 -*-
"""
Created on Thu Jul  8 18:57:53 2021

@author: MAmaro
"""

import os
os.environ["PROJ_LIB"] = r'C:\Users\mamaro\Anaconda3\Library\share\proj'
import networkx as nx
import datetime as dt
import pandas as pd
from pyvis.network import Network
from openpyxl import load_workbook
from mpl_toolkits.basemap import Basemap as Basemap
import numpy as np
path = "C:/Users/mamaro/OneDrive - Colbun S.A/jobs/2021/20.pdoTx/"
def setMW(anno, mes):
    fecha = dt.datetime(anno, mes, 1)
    fechaOperativa = str(anno) + str(mes).rjust(2, '0')
    pdoTxCap[fechaOperativa] = lineas['Operativa']*lineas['A->B']
    for linea in range(len(lineas)):
        for tabla in ['mantlin', 'mantlincsv']:
            df = mapping[tabla]
            # Se determina si la linea está operativa o no para la fecha fechaOperativa
            try:
                pdoTxCap.loc[linea, fechaOperativa] = \
                    ((df['Operativa']) *
                     (df['A->B']))\
                    .loc[(df['Línea'] == lineas['Nombre A->B'][linea]) &
                         (df['Inicial'] <= fecha) &
                         (df['Final'] >= fecha)].tail(1).item()
            except ValueError:
                pass
        if pdoTxCap[fechaOperativa][linea] == 0:
            txGraphx.add_edge(lineas['Barra A'][linea], lineas['Barra B'][linea], hidden = True)
        else:
            txGraphx.add_edge(lineas['Barra A'][linea], lineas['Barra B'][linea], hidden = False)

filename = path + "pdoTxIn-r3.xlsx"

# read file
wb = load_workbook(filename)

mapping = {}
# for x in range(len(wb.sheetnames)):
for ws in wb.worksheets:
    for entry, data_boundary in ws.tables.items():
        # print(entry)
        # parse the data within the ref boundary
        data = ws[data_boundary]
        # extract the data
        # the inner list comprehension gets the values for each cell in the table
        content = [[cell.value for cell in ent]
                   for ent in data
                   ]

        header = content[0]

        # the contents ... excluding the header
        rest = content[1:]

        # create dataframe with the column names
        # and pair table name with dataframe
        df = pd.DataFrame(rest, columns=header)
        mapping[entry] = df

barras = mapping['barras']
lineas = mapping['lineas']
mantlin = mapping['mantlin']
mantlincsv = mapping['mantlincsv']
mantlin['ini<fin'] = 1*(mantlin['Inicial'] < mantlin['Final'])
mantlincsv['ini<fin'] = 1*(mantlincsv['Inicial'] < mantlincsv['Final'])
mantlin = mantlin.loc[mantlin['ini<fin'] == 1]
mantlincsv = mantlincsv.loc[mantlincsv['ini<fin'] == 1]
# mantlincsv = mantlincsv.sort_values(['Línea', 'Inicial'])
mapping['mantlin'] = mantlin
mapping['mantlincsv'] = mantlincsv
txGraphx = nx.Graph()
m = Basemap(
        projection='merc',
        llcrnrlon=-84,
        llcrnrlat=-54,
        urcrnrlon=-64,
        urcrnrlat=-15,
        lat_ts=0,
        resolution='i',
        )
mx, my = m(barras['Longitud'], barras['Latitud'])
volts = ['66', '100', '110', '154', '220', '345', '500', '600']
barras['Voltaje'] = [int([volt for volt in volts if (volt in nom)][0]) for nom in barras['Barra']]
lineas['Trafo'] = [a != b for (a, b) in zip(
    barras['Voltaje'][lineas['Barra A'] - 1], 
    barras['Voltaje'][lineas['Barra B'] - 1]
    )]

v2c = {66: '#33FF3E', 100: '#EB8003', 110: '#EBA003', 154: '#33FFF3', 220: '#4333FF', 345: '#433388', 500: '#FF3333', 600: '#FFFFFF'}

node_data = zip(barras['N°'], barras['Barra'], np.array(mx)/200, -np.array(my)/1600, barras['Fijo'], barras['Voltaje'])
for n in node_data:
    num = n[0]
    nom = n[1]
    lon = n[2]
    lat = n[3]
    fij = n[4]
    col = v2c[n[5]]
    txGraphx.add_node(num, label=nom, x = lon, y = lat, size = 4, fixed = fij, color = col)

edge_data = zip(lineas['Barra A'], lineas['Barra B'], lineas['V [kV]'], lineas['Trafo'])
for e in edge_data:
    src = e[0]
    dst = e[1]
    col = v2c[e[2]]
    trf = e[3]
    txGraphx.add_edge(src, dst, color = col, width = 1, dashes = trf)

pdoTxCap = lineas['Nombre A->B'].to_frame()

setMW(2021, 7)
for anno in range(2022, 2029, 1):
    for mes in range(1, 13, 12):
        setMW(anno, mes)
        txGraph = Network(height='900px', width='100%', bgcolor='#222222', font_color='white')        
        txGraph.options = {
            "physics": {
                "repulsion": {
                    "centralGravity": 0,
                    "springLength": 5,
                    "springConstant": 0.02,
                    "nodeDistance": 50,
                    "damping": 0.5
                    },
                "maxVelocity": 150,
                "minVelocity": 0.75,
                "solver": "repulsion",
                "timestep": 1
                },
            "edges": {
                "arrows": {
                    "to": {
                        "enabled": True,
                        "scaleFactor": 0.5
                        }
                    }   ,
                "smooth": {
                    "type": "continuous",
                    "forceDirection": "none"
                    }
                } 
            }
        txGraph.from_nx(txGraphx)
        # nx.draw(txGraphx)
        # plt.savefig(path + 'txGraph'+ str(anno) + str(mes).rjust(2, '0') + '.pdf')
        txGraph.save_graph(path + 'txGraph'+ str(anno) + str(mes).rjust(2, '0') + '.html')

pdoTxCap = pdoTxCap.sort_values(pdoTxCap.columns[1:].tolist())

# Create a Pandas Excel writer using XlsxWriter as the engine.
writer = pd.ExcelWriter(path + 'pdoTxOut.xlsx', engine='xlsxwriter')

# Write the dataframe data to XlsxWriter. Turn off the default header and
# index and skip one row to allow us to insert a user defined header.
pdoTxCap.to_excel(writer, sheet_name='Sheet1',
               startrow=1, header=False, index=False)

# Get the xlsxwriter workbook and worksheet objects.
workbook = writer.book
worksheet = writer.sheets['Sheet1']

# Get the dimensions of the dataframe.
(max_row, max_col) = pdoTxCap.shape

# Create a list of column headers, to use in add_table().
column_settings = [{'header': column} for column in pdoTxCap.columns]

# Add the Excel table structure. Pandas will add the data.
worksheet.add_table(0, 0, max_row, max_col - 1, {'columns': column_settings})

# Make the columns wider for clarity.
worksheet.set_column(0, max_col - 1, 12)

# Close the Pandas Excel writer and output the Excel file.
writer.save()
writer.handles = None

